"""
Report Generation Utilities — Excel and PDF report builders.

Excel reports produce a multi-sheet professional workbook:
  Sheet 1 "Summary"  — executive cover + KPI table
  Sheet 2 "Data"     — full data with smart formatting
  Sheet 3 "Charts"   — native Excel charts from aggregated data
"""
import io
import os
import re
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from config import COLORS, CHART_COLORS


# ──────────────────────────────────────────────
# Brand palette (openpyxl hex — no '#' prefix)
# ──────────────────────────────────────────────
_NAVY       = "1B2A4A"
_STEEL_BLUE = "2E86AB"
_GOLD       = "D4AF37"
_LIGHT_GRAY = "F8F9FA"
_WHITE      = "FFFFFF"
_GREEN      = "28A745"
_RED        = "DC3545"
_AMBER      = "FFC107"
_DARK_NAVY  = "0D1B2A"
_CHART_HEX  = [c.lstrip("#") for c in CHART_COLORS]

# Style presets
_HEADER_FONT  = Font(name="Calibri", bold=True, color=_WHITE, size=11)
_HEADER_FILL  = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER  = Border(
    left=Side(style="thin", color="DEE2E6"),
    right=Side(style="thin", color="DEE2E6"),
    top=Side(style="thin", color="DEE2E6"),
    bottom=Side(style="thin", color="DEE2E6"),
)
_BAND_FILL    = PatternFill(start_color=_LIGHT_GRAY, end_color=_LIGHT_GRAY, fill_type="solid")
_NO_FILL      = PatternFill(fill_type=None)

# Conditional fills
_FILL_GREEN = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
_FILL_RED   = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
_FILL_AMBER = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

# Column type detection sets
_CURRENCY_COLS = {
    "balance", "income", "loan_amount", "amount", "outstanding_amount",
    "card_limit", "emi", "deposits", "total_deposits", "total_loan_amount",
    "avg_balance", "salary", "clv_score",
}
_PERCENT_COLS = {
    "churn_score", "churn_rate", "conversion_rate", "probability",
    "interest_rate",
}
_DATE_COLS = {
    "customer_since", "opened_date", "applied_date", "approved_date",
    "issued_date", "expiry_date", "date", "timestamp", "last_login",
    "created_at",
}
_STATUS_COLS = {"status", "risk_level", "is_active", "segment"}

# Conditional fill mapping for status values
_GREEN_VALUES = {"active", "low", "yes", "1", "closed", "premium", "high value"}
_RED_VALUES   = {"inactive", "high", "no", "0", "defaulted", "blocked", "expired", "dormant"}
_AMBER_VALUES = {"medium", "pending", "regular", "low value"}


# ──────────────────────────────────────────────
# Private Helpers
# ──────────────────────────────────────────────

def _col_key(col_name: str) -> str:
    """Normalize column name for type lookup."""
    return col_name.lower().replace(" ", "_")


def _to_title(col_name: str) -> str:
    """Convert snake_case column name to Title Case."""
    return col_name.replace("_", " ").title()


def _style_header_row(ws, num_cols: int):
    """Apply navy fill + white bold font + freeze panes to the header row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.freeze_panes = "A2"


def _get_conditional_fill(value) -> PatternFill | None:
    """Return a fill color based on status-like cell value."""
    if pd.isna(value):
        return None
    s = str(value).strip().lower()
    if s in _GREEN_VALUES:
        return _FILL_GREEN
    if s in _RED_VALUES:
        return _FILL_RED
    if s in _AMBER_VALUES:
        return _FILL_AMBER
    return None


def _style_data_cells(ws, df: pd.DataFrame, start_row: int = 2):
    """Apply number formats, conditional fills, borders, and row banding."""
    for col_idx, col_name in enumerate(df.columns, 1):
        key = _col_key(col_name)

        # Determine format
        num_fmt = None
        if key in _CURRENCY_COLS:
            num_fmt = '$#,##0.00'
        elif key in _PERCENT_COLS:
            num_fmt = '0.0%'
        elif key in _DATE_COLS:
            num_fmt = 'YYYY-MM-DD'

        is_status = key in _STATUS_COLS

        for row_idx in range(start_row, start_row + len(df)):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")

            # Row banding (alternate shading on even data rows)
            data_row = row_idx - start_row  # 0-indexed
            if data_row % 2 == 1:
                cell.fill = _BAND_FILL

            # Number format
            if num_fmt:
                cell.number_format = num_fmt

            # Conditional fill for status columns (overrides banding)
            if is_status:
                cond_fill = _get_conditional_fill(cell.value)
                if cond_fill:
                    cell.fill = cond_fill


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 40):
    """Auto-fit column widths based on max content length per column."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        # Add padding
        adjusted = min(max(max_len + 3, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _build_summary_sheet(wb: Workbook, report_title: str, user_name: str,
                         filters: dict, kpi_data: list[dict]):
    """
    Create Sheet 1 — executive cover section + KPI summary table.

    kpi_data: list of {"label": "...", "value": ...} dicts.
    """
    ws = wb.create_sheet("Summary", 0)

    # ── Title Block ──
    # Merge A1:F1 for app name banner
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "AI Banking Customer Insights"
    title_cell.font = Font(name="Calibri", bold=True, color=_WHITE, size=18)
    title_cell.fill = PatternFill(start_color=_DARK_NAVY, end_color=_NAVY, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 45

    # Subtitle row — report title
    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = report_title
    sub.font = Font(name="Calibri", bold=True, color=_GOLD, size=14)
    sub.fill = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 32

    # Metadata rows
    meta_start = 4
    meta_items = [
        ("Generated By", user_name),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    if filters:
        for k, v in filters.items():
            if v:
                meta_items.append((k, str(v)))

    for i, (label, value) in enumerate(meta_items):
        row = meta_start + i
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, color=_NAVY, size=11)
        ws.cell(row=row, column=2, value=value).font = Font(name="Calibri", size=11)

    # ── KPI Summary Table ──
    kpi_start = meta_start + len(meta_items) + 2
    ws.cell(row=kpi_start, column=1, value="Key Performance Indicators").font = Font(
        name="Calibri", bold=True, color=_NAVY, size=13
    )
    kpi_start += 1

    # Headers
    for col_idx, header in enumerate(["Metric", "Value"], 1):
        cell = ws.cell(row=kpi_start, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # KPI rows
    for i, kpi in enumerate(kpi_data):
        row = kpi_start + 1 + i
        ws.cell(row=row, column=1, value=kpi["label"]).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=row, column=1).border = _THIN_BORDER
        val_cell = ws.cell(row=row, column=2, value=kpi["value"])
        val_cell.font = Font(name="Calibri", size=11)
        val_cell.border = _THIN_BORDER
        val_cell.alignment = Alignment(horizontal="right")
        # Apply format
        if isinstance(kpi["value"], (int, float)):
            if kpi.get("format") == "currency":
                val_cell.number_format = '$#,##0.00'
            elif kpi.get("format") == "percent":
                val_cell.number_format = '0.0%'
            elif kpi.get("format") == "integer":
                val_cell.number_format = '#,##0'
        # Alternate banding
        if i % 2 == 1:
            ws.cell(row=row, column=1).fill = _BAND_FILL
            ws.cell(row=row, column=2).fill = _BAND_FILL

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    for col_letter in ["C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 15

    # Print setup
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _build_data_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str = "Data"):
    """Create Sheet 2 — full data table with professional formatting."""
    ws = wb.create_sheet(sheet_name)

    # Write Title Case headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=_to_title(col_name))

    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(df.columns, 1):
            val = row[col_name]
            # Convert numpy types
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            elif pd.isna(val):
                val = None
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Apply styling
    num_cols = len(df.columns)
    _style_header_row(ws, num_cols)
    _style_data_cells(ws, df, start_row=2)
    _auto_fit_columns(ws)

    # Print setup
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:{get_column_letter(num_cols)}{len(df) + 1}"


def _write_chart_data(ws, labels: list, values: list, start_row: int = 1,
                       label_col: int = 1, value_col: int = 2,
                       header_label: str = "Category", header_value: str = "Value"):
    """Write a label/value series to a worksheet for chart binding."""
    ws.cell(row=start_row, column=label_col, value=header_label)
    ws.cell(row=start_row, column=value_col, value=header_value)
    for i, (lbl, val) in enumerate(zip(labels, values)):
        ws.cell(row=start_row + 1 + i, column=label_col, value=str(lbl))
        ws.cell(row=start_row + 1 + i, column=value_col, value=float(val) if val is not None else 0)
    return start_row, start_row + len(labels)


def _make_pie_chart(ws_data, title: str, start_row: int, end_row: int,
                    label_col: int = 1, value_col: int = 2) -> PieChart:
    """Create a styled PieChart from data range."""
    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data_ref = Reference(ws_data, min_col=value_col, min_row=start_row,
                         max_row=end_row)
    cats_ref = Reference(ws_data, min_col=label_col, min_row=start_row + 1,
                         max_row=end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Apply brand colors
    series = chart.series[0]
    for i in range(end_row - start_row):
        pt = DataPoint(idx=i)
        color_hex = _CHART_HEX[i % len(_CHART_HEX)]
        pt.graphicalProperties.solidFill = color_hex
        series.data_points.append(pt)

    chart.legend.position = 'b'
    # Data labels
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal = False
    return chart


def _make_bar_chart(ws_data, title: str, start_row: int, end_row: int,
                    label_col: int = 1, value_col: int = 2,
                    x_title: str = "", y_title: str = "") -> BarChart:
    """Create a styled BarChart from data range."""
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data_ref = Reference(ws_data, min_col=value_col, min_row=start_row,
                         max_row=end_row)
    cats_ref = Reference(ws_data, min_col=label_col, min_row=start_row + 1,
                         max_row=end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    if x_title:
        chart.x_axis.title = x_title
    if y_title:
        chart.y_axis.title = y_title

    # Brand color
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = _CHART_HEX[0]

    chart.legend = None
    return chart


def _make_line_chart(ws_data, title: str, start_row: int, end_row: int,
                     label_col: int = 1, value_col: int = 2,
                     x_title: str = "", y_title: str = "") -> LineChart:
    """Create a styled LineChart from data range."""
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data_ref = Reference(ws_data, min_col=value_col, min_row=start_row,
                         max_row=end_row)
    cats_ref = Reference(ws_data, min_col=label_col, min_row=start_row + 1,
                         max_row=end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    if x_title:
        chart.x_axis.title = x_title
    if y_title:
        chart.y_axis.title = y_title

    # Brand color + smooth
    if chart.series:
        chart.series[0].graphicalProperties.line.solidFill = _CHART_HEX[0]
        chart.series[0].smooth = True

    chart.legend = None
    return chart


def _build_charts_sheet(wb: Workbook, df: pd.DataFrame, report_type: str):
    """
    Create Sheet 3 — native Excel charts.

    Writes aggregated data to a hidden helper sheet '_ChartData',
    then builds chart objects on a visible 'Charts' sheet.
    """
    if report_type == "Executive Dashboard Summary":
        return  # No charts for executive summary

    ws_charts = wb.create_sheet("Charts")

    if df.empty:
        ws_charts["B2"] = "No data matching the filter criteria. Charts cannot be generated."
        ws_charts["B2"].font = Font(name="Calibri", italic=True, color="6C757D", size=11)
        return

    ws_data = wb.create_sheet("_ChartData")
    ws_data.sheet_state = "hidden"

    current_data_row = 1
    chart_anchors = ["B2", "B20", "B38"]
    chart_idx = 0

    def _add(chart_obj):
        nonlocal chart_idx
        if chart_idx < len(chart_anchors):
            ws_charts.add_chart(chart_obj, chart_anchors[chart_idx])
            chart_idx += 1

    # ── Customer Report ──
    if report_type in ("Customer Report", "Customer Segmentation Report"):
        # Chart 1: Region distribution
        if "region" in df.columns:
            counts = df["region"].value_counts()
            start, end = _write_chart_data(
                ws_data, counts.index.tolist(), counts.values.tolist(),
                start_row=current_data_row, header_value="Customers"
            )
            _add(_make_pie_chart(ws_data, "Customer Distribution by Region",
                                 current_data_row, end + 1))
            current_data_row = end + 3

        # Chart 2: Segment distribution
        if "segment" in df.columns:
            counts = df["segment"].value_counts()
            start, end = _write_chart_data(
                ws_data, counts.index.tolist(), counts.values.tolist(),
                start_row=current_data_row, header_value="Customers"
            )
            _add(_make_bar_chart(ws_data, "Customer Segments",
                                 current_data_row, end + 1,
                                 x_title="Segment", y_title="Count"))
            current_data_row = end + 3

        # Chart 3: Credit score distribution (histogram-style buckets)
        if "credit_score" in df.columns:
            bins = [300, 400, 500, 600, 700, 800, 850]
            labels_hist = ["300-399", "400-499", "500-599", "600-699", "700-799", "800-850"]
            hist_counts = pd.cut(df["credit_score"], bins=bins, labels=labels_hist, right=False).value_counts().sort_index()
            start, end = _write_chart_data(
                ws_data, hist_counts.index.tolist(), hist_counts.values.tolist(),
                start_row=current_data_row, header_label="Score Range", header_value="Customers"
            )
            _add(_make_bar_chart(ws_data, "Credit Score Distribution",
                                 current_data_row, end + 1,
                                 x_title="Credit Score Range", y_title="Count"))
            current_data_row = end + 3

    # ── Loan Report ──
    elif report_type == "Loan Report":
        # Chart 1: Loan type distribution
        if "loan_type" in df.columns:
            counts = df["loan_type"].value_counts()
            start, end = _write_chart_data(
                ws_data, counts.index.tolist(), counts.values.tolist(),
                start_row=current_data_row, header_value="Loans"
            )
            _add(_make_pie_chart(ws_data, "Loan Type Distribution",
                                 current_data_row, end + 1))
            current_data_row = end + 3

        # Chart 2: Loan status
        if "status" in df.columns:
            counts = df["status"].value_counts()
            start, end = _write_chart_data(
                ws_data, counts.index.tolist(), counts.values.tolist(),
                start_row=current_data_row, header_value="Loans"
            )
            _add(_make_bar_chart(ws_data, "Loan Status Distribution",
                                 current_data_row, end + 1,
                                 x_title="Status", y_title="Count"))
            current_data_row = end + 3

        # Chart 3: Average loan amount by type
        if "loan_type" in df.columns and "loan_amount" in df.columns:
            avg_by_type = df.groupby("loan_type")["loan_amount"].mean().round(2)
            start, end = _write_chart_data(
                ws_data, avg_by_type.index.tolist(), avg_by_type.values.tolist(),
                start_row=current_data_row, header_label="Loan Type", header_value="Avg Amount"
            )
            _add(_make_bar_chart(ws_data, "Average Loan Amount by Type",
                                 current_data_row, end + 1,
                                 x_title="Loan Type", y_title="Amount ($)"))
            current_data_row = end + 3

    # ── Transaction Report ──
    elif report_type == "Transaction Report":
        # Chart 1: Transaction type distribution
        if "type" in df.columns:
            counts = df["type"].value_counts()
            start, end = _write_chart_data(
                ws_data, counts.index.tolist(), counts.values.tolist(),
                start_row=current_data_row, header_value="Transactions"
            )
            _add(_make_pie_chart(ws_data, "Transaction Type Distribution",
                                 current_data_row, end + 1))
            current_data_row = end + 3

        # Chart 2: Monthly volume (if date column exists)
        if "date" in df.columns:
            try:
                df_temp = df.copy()
                df_temp["month"] = pd.to_datetime(df_temp["date"], errors="coerce").dt.to_period("M").astype(str)
                monthly = df_temp.groupby("month").size()
                if len(monthly) > 1:
                    start, end = _write_chart_data(
                        ws_data, monthly.index.tolist(), monthly.values.tolist(),
                        start_row=current_data_row, header_label="Month", header_value="Volume"
                    )
                    _add(_make_line_chart(ws_data, "Monthly Transaction Volume",
                                          current_data_row, end + 1,
                                          x_title="Month", y_title="Transactions"))
                    current_data_row = end + 3
            except Exception:
                pass

        # Chart 3: Channel distribution
        if "channel" in df.columns:
            counts = df["channel"].value_counts()
            start, end = _write_chart_data(
                ws_data, counts.index.tolist(), counts.values.tolist(),
                start_row=current_data_row, header_value="Transactions"
            )
            _add(_make_bar_chart(ws_data, "Transaction Channel Distribution",
                                 current_data_row, end + 1,
                                 x_title="Channel", y_title="Count"))
            current_data_row = end + 3

    # ── Churn Report ──
    elif report_type == "Churn Report":
        # Chart 1: Risk level distribution
        if "risk_level" in df.columns:
            counts = df["risk_level"].value_counts()
            start, end = _write_chart_data(
                ws_data, counts.index.tolist(), counts.values.tolist(),
                start_row=current_data_row, header_value="Customers"
            )
            _add(_make_pie_chart(ws_data, "Risk Level Distribution",
                                 current_data_row, end + 1))
            current_data_row = end + 3

        # Chart 2: Churn score histogram
        if "churn_score" in df.columns:
            bins = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
            labels_hist = ["0-10%", "10-20%", "20-30%", "30-40%", "40-50%",
                           "50-60%", "60-70%", "70-80%", "80-90%", "90-100%"]
            hist_counts = pd.cut(df["churn_score"], bins=bins, labels=labels_hist, right=False).value_counts().sort_index()
            start, end = _write_chart_data(
                ws_data, hist_counts.index.tolist(), hist_counts.values.tolist(),
                start_row=current_data_row, header_label="Churn Score", header_value="Customers"
            )
            _add(_make_bar_chart(ws_data, "Churn Score Distribution",
                                 current_data_row, end + 1,
                                 x_title="Churn Score Range", y_title="Count"))
            current_data_row = end + 3

    # Title for the Charts sheet
    ws_charts["B1"] = f"Charts — {report_type}"
    ws_charts["B1"].font = Font(name="Calibri", bold=True, color=_NAVY, size=14)


# ──────────────────────────────────────────────
# Public API — Excel
# ──────────────────────────────────────────────

def generate_excel_report(report_title: str, data: pd.DataFrame,
                          metadata: dict = None, sheet_name: str = "Report",
                          user_name: str = "System", filters: dict = None,
                          kpi_data: list[dict] = None,
                          report_type: str = "",
                          include_summary: bool = True,
                          include_data: bool = True,
                          include_charts: bool = True) -> io.BytesIO:
    """
    Generate a professional multi-sheet Excel report.

    Args:
        report_title: Display title for the report.
        data: Main DataFrame to export.
        metadata: Legacy metadata dict (still supported for backward compat).
        sheet_name: Name for the data sheet.
        user_name: Name of the user generating the report.
        filters: Dict of filter labels/values to display on the Summary sheet.
        kpi_data: List of {"label", "value", "format"} dicts for KPIs.
        report_type: One of the REPORT_TYPES keys (drives chart selection).
        include_summary: Whether to include the Summary sheet.
        include_data: Whether to include the Data sheet.
        include_charts: Whether to include the Charts sheet.

    Returns:
        BytesIO buffer containing the .xlsx file.
    """
    wb = Workbook()
    # Remove the default sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Sheet 1 — Summary
    if include_summary:
        _build_summary_sheet(wb, report_title, user_name, filters or {},
                             kpi_data or [])

    # Sheet 2 — Data
    if include_data:
        _build_data_sheet(wb, data, sheet_name="Data")

    # Sheet 3 — Charts
    if include_charts and report_type != "Executive Dashboard Summary":
        _build_charts_sheet(wb, data, report_type)

    # Ensure at least one visible sheet
    if not wb.sheetnames:
        _build_data_sheet(wb, data, sheet_name="Data")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ──────────────────────────────────────────────
# Public API — PDF (unchanged)
# ──────────────────────────────────────────────

def generate_pdf_report(report_title: str, data: pd.DataFrame,
                        metadata: dict = None, max_rows: int = 100) -> io.BytesIO:
    """
    Generate a formatted PDF report with banking theme.

    Args:
        report_title: Title for the report
        data: DataFrame with report data
        metadata: Optional dict of metadata
        max_rows: Maximum rows to include in the PDF table

    Returns:
        BytesIO buffer containing the PDF file
    """
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                           topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Title'],
        fontSize=18, textColor=colors.HexColor('#1B2A4A'),
        spaceAfter=12
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#6C757D'),
        spaceAfter=20
    )

    # Title
    elements.append(Paragraph(f"🏦 {report_title}", title_style))
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    elements.append(Paragraph(f"Generated: {generated_at}", subtitle_style))

    # Metadata
    if metadata:
        meta_data = [[str(k), str(v)] for k, v in metadata.items()]
        meta_table = Table(meta_data, colWidths=[2*inch, 4*inch])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1B2A4A')),
        ]))
        elements.append(meta_table)
        elements.append(Spacer(1, 0.3*inch))

    # Data table
    display_cols = data.columns[:8].tolist()  # Limit columns for readability
    table_data = [display_cols]
    for _, row in data.head(max_rows).iterrows():
        table_data.append([str(row[col])[:25] for col in display_cols])

    if table_data:
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B2A4A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#D4AF37')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DEE2E6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ]))
        elements.append(t)

    # Footer
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph(
        f"AI Banking Customer Insights Platform · {len(data)} records · Page 1",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ──────────────────────────────────────────────
# Data Fetching with Filter Support
# ──────────────────────────────────────────────

def get_report_data(report_type: str, conn, filters: dict = None) -> pd.DataFrame:
    """
    Fetch data for a specific report type with optional SQL-level filters.

    Args:
        report_type: Type of report to generate
        conn: Database connection
        filters: Optional dict with keys: region, branch, date_from, date_to

    Returns:
        DataFrame with report data
    """
    base_queries = {
        "Customer Report": "SELECT * FROM customers",
        "Loan Report": "SELECT * FROM loans",
        "Transaction Report": "SELECT * FROM transactions",
        "Customer Segmentation Report": "SELECT * FROM customers",
        "Churn Report": "SELECT * FROM customers WHERE is_active = 0",
    }

    query = base_queries.get(report_type)
    if query is None:
        return pd.DataFrame()

    params = []
    where_clauses = []

    # Determine if this is a customer-based or other table
    is_customer_table = report_type in (
        "Customer Report", "Customer Segmentation Report", "Churn Report"
    )

    if filters:
        # Region filter (customer tables only)
        if is_customer_table and filters.get("region"):
            regions = filters["region"]
            if isinstance(regions, list) and regions:
                placeholders = ",".join(["?"] * len(regions))
                where_clauses.append(f"region IN ({placeholders})")
                params.extend(regions)

        # Branch filter (customer tables only)
        if is_customer_table and filters.get("branch"):
            branches = filters["branch"]
            if isinstance(branches, list) and branches:
                placeholders = ",".join(["?"] * len(branches))
                where_clauses.append(f"branch IN ({placeholders})")
                params.extend(branches)

        # Date range filter
        date_from = filters.get("date_from")
        date_to = filters.get("date_to")
        if date_from or date_to:
            if report_type == "Transaction Report":
                date_col = "date"
            elif report_type == "Loan Report":
                date_col = "applied_date"
            elif is_customer_table:
                date_col = "customer_since"
            else:
                date_col = None

            if date_col:
                if date_from:
                    where_clauses.append(f"{date_col} >= ?")
                    params.append(str(date_from))
                if date_to:
                    where_clauses.append(f"{date_col} <= ?")
                    params.append(str(date_to))

    # Build final query
    if where_clauses:
        # If query already has WHERE, use AND; otherwise add WHERE
        if " WHERE " in query.upper():
            query += " AND " + " AND ".join(where_clauses)
        else:
            query += " WHERE " + " AND ".join(where_clauses)

    return pd.read_sql(query, conn, params=params)


def get_kpi_data(conn, report_type: str = "") -> list[dict]:
    """
    Compute KPI metrics from the database for the Summary sheet.

    Returns list of {"label", "value", "format"} dicts.
    """
    kpis = []

    try:
        total_customers = pd.read_sql(
            "SELECT COUNT(*) as c FROM customers", conn
        ).iloc[0]["c"]
        active_customers = pd.read_sql(
            "SELECT COUNT(*) as c FROM customers WHERE is_active = 1", conn
        ).iloc[0]["c"]
        avg_balance = pd.read_sql(
            "SELECT COALESCE(AVG(balance), 0) as a FROM customers", conn
        ).iloc[0]["a"]
        total_deposits = pd.read_sql(
            "SELECT COALESCE(SUM(amount), 0) as s FROM transactions WHERE type='Deposit'",
            conn
        ).iloc[0]["s"]
        total_loans = pd.read_sql(
            "SELECT COALESCE(SUM(loan_amount), 0) as s FROM loans", conn
        ).iloc[0]["s"]
        total_accounts = pd.read_sql(
            "SELECT COUNT(*) as c FROM accounts", conn
        ).iloc[0]["c"]
        total_txns = pd.read_sql(
            "SELECT COUNT(*) as c FROM transactions", conn
        ).iloc[0]["c"]

        churn_rate = 1 - (active_customers / total_customers) if total_customers > 0 else 0

        kpis = [
            {"label": "Total Customers", "value": total_customers, "format": "integer"},
            {"label": "Active Customers", "value": active_customers, "format": "integer"},
            {"label": "Total Accounts", "value": total_accounts, "format": "integer"},
            {"label": "Total Transactions", "value": total_txns, "format": "integer"},
            {"label": "Average Balance", "value": round(avg_balance, 2), "format": "currency"},
            {"label": "Total Deposits", "value": round(total_deposits, 2), "format": "currency"},
            {"label": "Total Loan Portfolio", "value": round(total_loans, 2), "format": "currency"},
            {"label": "Churn Rate", "value": round(churn_rate, 4), "format": "percent"},
        ]
    except Exception:
        kpis = [{"label": "Error", "value": "Unable to compute KPIs", "format": "text"}]

    return kpis
